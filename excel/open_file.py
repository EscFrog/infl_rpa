from openpyxl import load_workbook # 파일 불렁기 모듈
wb = load_workbook("sample_cell.xlsx") # sample.xlsx 파일을 불러와서 wb에 할당
ws = wb.active # 활성화된 Sheet

# cell 데이터 불러오기
for x in range(5, 15):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()