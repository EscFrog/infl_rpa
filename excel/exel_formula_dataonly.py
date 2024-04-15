from openpyxl import load_workbook

'''
wb = load_workbook("sample_formular.xlsx")
ws = wb.active

# 수식 그대로 가져오고 있음
for row in ws.values:
    for cell in row:
        print(cell)
'''     
     
wb = load_workbook("sample_formular.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않는 상태의 데이터는 None 이라고 표시.
# openpyxl로 만든 파일은 직접 한 번 열었다가 저장해줘야 evaluate가 됨.
for row in ws.values:
    for cell in row:
        print(cell)

