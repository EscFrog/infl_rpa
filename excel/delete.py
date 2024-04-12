from openpyxl import load_workbook
wb = load_workbook("sample_cell_range.xlsx")
ws = wb.active

ws.delete_rows(8, 3) # 8번째 행부터 3개 행 삭제
wb.save("sample_del_row.xlsx")


wb = load_workbook("sample_cell_range.xlsx")
ws = wb.active

ws.delete_cols(2, 2) # 1번째 열부터 2개 열 삭제
wb.save("sample_del_col.xlsx")