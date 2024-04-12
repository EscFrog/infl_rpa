from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import random

wb = load_workbook("sample_cell_range.xlsx")
ws = wb.active

# 제목을 입력 받아 해당 제목의 열이 몇 번째인지 반환한다.
def search_title_col_num(title):
    col_num = 0
    for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        if col[0].value == title:
            col_num = idx
            break
    return col_num

# 새로운 "과학" 열 추가
new_col_letter = get_column_letter(ws.max_column + 1)
ws[f"{new_col_letter}1"] = "과학"
for row_num in range(2, ws.max_row + 1):
    ws[f"{new_col_letter}{row_num}"] = random.randint(0, 100)

# 데이터 범위를 한 열씩 오른쪽으로 이동하고 "국어"열 추가
last_column_letter = get_column_letter(ws.max_column)
range_to_move = f"B1:{last_column_letter}{ws.max_row}"
ws.move_range(range_to_move, rows=0, cols=1)
ws["B1"].value = "국어"
for row_num in range(2, ws.max_row + 1):
    ws[f"B{row_num}"] = random.randint(0, 100)

# '영어' 열과 '수학' 열 찾기
eng_col_num = search_title_col_num("영어")
math_col_num = search_title_col_num("수학")
math_col_letter = get_column_letter(math_col_num)

# '수학' 열 전체를 잘라서 '영어' 열 아래에 붙이기
range_to_move = f"{math_col_letter}1:{math_col_letter}{ws.max_row}"
ws.move_range(range_to_move, rows=ws.max_row, cols=eng_col_num - math_col_num)

wb.save("sample_move.xlsx")