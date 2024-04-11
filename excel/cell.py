from openpyxl import Workbook
# from random import *

import random

wb = Workbook()
ws = wb.active
ws.title = "escFrog Sheet"

# A1 셀에 1 이라는 값을 입력
ws["a1"] = 1 # 소문자로 해도 상관없음
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 값을 출력
print(ws["A10"].value) # 값이 없을 땐 "None" 출력

print(ws.cell(row=1, column=1).value) # ws["A1"].value 와 동일
print(ws.cell(column=2, row=1).value) # ws["B1"].value 와 동일

c = ws.cell(column=4, row=1, value="some string") # 이런식으로 셀에 값을 입력할 수도 있음
print(c.value) # ws["c1"]

# 반복문을 이용해서 랜덤 숫자 채우기
for x in range(5, 15):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=random.randint(0, 100))

# 숫자 순서대로 채우기
number = 1
for x in range(16, 26):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=number)
        number += 1


wb.save("sample_cell.xlsx")