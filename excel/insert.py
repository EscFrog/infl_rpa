from openpyxl import load_workbook
wb = load_workbook("sample_cell_range.xlsx")
ws = wb.active

ws.insert_rows(8) # 8번째 행에 새 행을 넣고 나머지 행을 아래로 밈
ws.insert_rows(8, 5) # 8번째 행부터 5행을 넣고 나머지 행을 아래로 밈
wb.save("sample_insert_rows.xlsx")

# 새로 열자
wb = load_workbook("sample_cell_range.xlsx")
ws = wb.active

ws.insert_cols(2) # 2번째 열에 새 열을 넣고 나머지 열을 오른쪽으로 밈
ws.insert_cols(2, 3) # 2번째 열부터 3개의 열을 넣고 나머지 열을 오른쪽으로 밈
wb.save("sample_insert_cols.xlsx")