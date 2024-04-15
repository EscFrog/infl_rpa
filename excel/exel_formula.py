import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["a1"] = datetime.datetime.today() # 오늘 날짜 정버
ws.column_dimensions["a"].width = 20

ws["a2"] = "=SUM(1, 2, 3)"
ws["a3"] = "=AVERAGE(1, 2, 3)"

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

wb.save("sample_formular.xlsx")