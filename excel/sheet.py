from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() # 새로운 Sheet를 기본 이름으로 생성
ws.title = "MySheet" # 시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # 탭 색상을 RGB값의 색상으로 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 Sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근
print(new_ws.title)

print(wb.sheetnames)

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"


wb.save("sample.xlsx")