from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
import random

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, random.randint(0, 100), random.randint(0, 100)])

'''
col_B = ws["B"] # 영어 column 만 가져오기
# print(col_B)
for cell in col_B:
    print(cell.value)

col_range = ws["b:c"] # 영어 수학 column 함께 가져오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1] # 1번째 row만 가지고 오기
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가져오기
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        print(cell.coordinate, "=>", cell.value, end=" <= ")
        xy = coordinate_from_string(cell.coordinate)
        print(f"${xy[0]}", end="")
        print(f"${xy[1]}", end=" // ")
    print()
'''

'''
# 전체 rows
print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[1].value)

# 전체 columns
print(tuple(ws.columns))
for column in tuple(ws.columns):
    print(column[1].value)
'''

'''
for idx, row in enumerate(ws.iter_rows()): # 전체 row에 대해 반복
    print(f"{idx}열:", end=" ")
    for cell in row:
        print(f"{cell.value}", end=" ")
    print()

for column in ws.iter_cols(): # 전체 column에 대해 반복
    print(column[0].value)
'''
 
# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value) # 수학, 영어 점수

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save("sample_cell_range.xlsx")