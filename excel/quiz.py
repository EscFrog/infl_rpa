from openpyxl import Workbook

wb = Workbook()
ws = wb.active

data = [
    [1,10,8,5,14,26,12],
    [2,7,3,7,15,24,18],
    [3,9,5,8,8,12,4],
    [4,7,8,7,17,21,18],
    [5,7,8,7,16,25,15],
    [6,3,5,8,8,17,0],
    [7,4,9,10,16,27,18],
    [8,6,6,6,15,19,17],
    [9,10,10,9,19,30,19],
    [10,9,8,8,20,25,20],
    ]

# 제목 행 생성
ws["a1"] = "학번"
ws["b1"] = "출석"
ws["c1"] = "퀴즈1"
ws["d1"] = "퀴즈2"
ws["e1"] = "중간고사"
ws["f1"] = "기말고사"
ws["g1"] = "프로젝트"

# 셀에 데이터 입력
for row_idx, row in enumerate(data):
    for col_idx, cell in enumerate(row):
        ws.cell(row=(row_idx + 2), column=(col_idx + 1), value=cell)

# 퀴즈2 점수를 모두 10으로 수정
for i in range(0, len(data)):
    data[i][3] = 10
    ws[f"D{i+2}"] = 10

# H열에 총점 추가
ws["H1"] = "총점"
for i in range(0, len(data)):
    row_num = i + 2
    ws[f"H{row_num}"] = f"=SUM(B{row_num}:G{row_num})"

# I열에 성적 정보 추가
ws["I1"] = "성적"
for row_idx, row in enumerate(data):
    
    total_point = sum(row) - row[0]
    
    grades = "D"
    
    # 총점에 따라 성적 매기기
    if total_point >= 90:
        grades = "A"
    elif total_point >= 80:
        grades = "B"
    elif total_point >= 70:
        grades = "C"
    
    # 출석점수가 5점 미만이면 F
    if row[1] < 5:
        grades = "F"
    
    print(sum(row), total_point, grades)
    ws[f"I{row_idx + 2}"] = grades
        

wb.save("scores.xlsx")